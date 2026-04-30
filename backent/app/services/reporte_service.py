import io
import csv
from typing import List, Any, Dict

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from openpyxl import Workbook


def generar_pdf(titulo: str, columnas: List[str], datos: List[List[Any]]) -> io.BytesIO:
    """Genera un reporte en PDF y lo retorna como BytesIO."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    styles = getSampleStyleSheet()

    import os
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    logo_path = os.path.join(os.path.dirname(BASE_DIR), "frontend", "assets", "logo.png")
    
    try:
        logo = Image(logo_path, width=80, height=80)
    except Exception:
        logo = Paragraph("<i>[Logo]</i>", styles["Normal"])

    title_p = Paragraph(f"<b>Binglish - {titulo}</b>", styles["Title"])
    
    header_table = Table([[logo, title_p]], colWidths=[1.0 * inch, 8.0 * inch])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "LEFT"),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Tabla
    table_data = [columnas] + datos
    table = Table(table_data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ECF0F1")),
        ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#BDC3C7")),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TOPPADDING", (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_excel(titulo: str, columnas: List[str], datos: List[List[Any]]) -> io.BytesIO:
    """Genera un reporte en Excel y lo retorna como BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita a 31 caracteres

    # Encabezados
    ws.append(columnas)

    # Estilos de encabezado
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row in datos:
        ws.append(row)

    # Ajustar ancho de columnas
    for i, col in enumerate(columnas, 1):
        max_len = max(len(str(col)), *(len(str(row[i - 1])) for row in datos) if datos else [0])
        ws.column_dimensions[chr(64 + i)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_csv(columnas: List[str], datos: List[List[Any]]) -> io.StringIO:
    """Genera un reporte en CSV y lo retorna como StringIO."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columnas)
    writer.writerows(datos)
    buffer.seek(0)
    return buffer
